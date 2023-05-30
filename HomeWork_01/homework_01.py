# Read data from xlsx file
import openpyxl

debt_value = input('Podaj wysokość zaciągniętego kredytu: ')
interest_rate = input('Podaj oprocentowanie kredytu w %: ')
due_monthly = input('Podaj wysokość miesięcznej raty kredytu: ')

# Check if "," was used instead of "." as decimal point and change string into float
if ',' in debt_value:
    debt_value = debt_value.replace(',', '.')
debt_value = float(debt_value)
if ',' in interest_rate:
    interest_rate = interest_rate.replace(',', '.')
#percent_sign = False
if '%' in interest_rate:
    interest_rate = interest_rate.replace('%', '').strip()
#    percent_sign = True
interest_rate = float(interest_rate)/100
#if percent_sign:
#    interest_rate = interest_rate / 100
if ',' in due_monthly:
    due_monthly = due_monthly.replace(',', '.')
due_monthly = float(due_monthly)

# Define variable to load the Excel workbook
dataframe = openpyxl.load_workbook("Oprocentowanie_pozyczki_oszczednosci_na_koncie.xlsx")

# Define variable to read active sheet
datasheet = dataframe.active  # since there's only one worksheet in the xlsx file it will always be the active one

print('\nPozostałe zadłużenie w kolejnych miesiącach: ')
# Iterate the loop to read the cell values and calculate remaining debt
for row in range(3, datasheet.max_row+1):
    month = datasheet.cell(row, 1).value
    inflation = datasheet.cell(row, 2).value/100#.replace(',', '.')
    #inflation = float(inflation)/100
    debt_value = debt_value*(1+interest_rate/12+inflation/12) - due_monthly
    print(f'{month} : {debt_value}')
